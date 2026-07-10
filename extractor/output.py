"""JSON and Excel writers for extracted invoices."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .schema import Invoice

HEADER_FONT = Font(bold=True)


def write_json(inv: Invoice, out_path: Path) -> None:
    out_path.write_text(json.dumps(inv.model_dump(), indent=2) + "\n")


def write_excel(invoices: dict[str, Invoice], out_path: Path) -> None:
    """One summary sheet, one combined line-items sheet."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Invoices"
    headers = ["File", "Vendor", "Invoice #", "Date", "PO #", "Subtotal", "Tax", "Total", "Warnings"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for fname, inv in invoices.items():
        ws.append([
            fname, inv.vendor_name, inv.invoice_number, inv.invoice_date,
            inv.po_number or "", inv.subtotal, inv.tax, inv.total,
            "; ".join(inv.warnings),
        ])

    ws2 = wb.create_sheet("Line items")
    ws2.append(["File", "Invoice #", "Description", "Qty", "Unit price", "Amount"])
    for cell in ws2[1]:
        cell.font = HEADER_FONT
    for fname, inv in invoices.items():
        for li in inv.line_items:
            ws2.append([fname, inv.invoice_number, li.description, li.quantity, li.unit_price, li.amount])

    for sheet in (ws, ws2):
        for col in sheet.columns:
            width = max(len(str(c.value or "")) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = min(width, 50)

    wb.save(out_path)
